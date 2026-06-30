import os
from datetime import datetime, date
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user,
    login_required, current_user
)
from werkzeug.security import generate_password_hash, check_password_hash
import csv
import io
from openpyxl import load_workbook, Workbook
from pdf_generator import generate_pi_pdf

basedir = os.path.abspath(os.path.dirname(__file__))
os.makedirs(os.path.join(basedir, 'instance'), exist_ok=True)

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-this-secret-key-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL', f"sqlite:///{os.path.join(basedir, 'instance', 'app.db')}"
).replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = "Please log in to access this page."

GST_RATE = 0.18

DISPATCH_UNITS = ["UNIT 1", "DABASPETE", "MAHIMAPURA", "DADRA"]
UOM_OPTIONS = ["Nos", "KG", "Pieces", "BAG", "TON", "BOX"]

# ---------------------------------------------------------------------------
# MODELS
# ---------------------------------------------------------------------------

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='staff')  # 'admin' or 'staff'
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, raw_password):
        self.password_hash = generate_password_hash(raw_password)

    def check_password(self, raw_password):
        return check_password_hash(self.password_hash, raw_password)

    def is_admin(self):
        return self.role == 'admin'


class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    contact_person = db.Column(db.String(120))
    phone = db.Column(db.String(30))
    email = db.Column(db.String(120))
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    created_by = db.relationship('User', backref='customers')


class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    default_uom = db.Column(db.String(20))
    default_price = db.Column(db.Float, default=0)
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    created_by = db.relationship('User', backref='items')


class Document(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    doc_type = db.Column(db.String(10), default='PI')  # PI or Quote
    quote_no = db.Column(db.String(20), unique=True, nullable=False)
    doc_date = db.Column(db.Date, default=date.today)
    dispatch_from = db.Column(db.String(50))

    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    item_desc = db.Column(db.String(300), nullable=False)
    packaging = db.Column(db.String(50))
    qty = db.Column(db.Float, default=0)
    uom = db.Column(db.String(20))
    price = db.Column(db.Float, default=0)

    base_amount = db.Column(db.Float, default=0)
    gst_applied = db.Column(db.Boolean, default=False)
    gst_amount = db.Column(db.Float, default=0)
    freight_charges = db.Column(db.Float, default=0)
    total_amount = db.Column(db.Float, default=0)

    follow_up_date = db.Column(db.Date)
    status = db.Column(db.String(20), default='pending')  # pending / delivered
    notes = db.Column(db.Text)

    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    customer = db.relationship('Customer', backref='documents')
    created_by = db.relationship('User', backref='documents')

    def recalc(self):
        self.base_amount = round((self.qty or 0) * (self.price or 0), 2)
        self.gst_amount = round(self.base_amount * GST_RATE, 2) if self.gst_applied else 0
        self.total_amount = round(self.base_amount + self.gst_amount + (self.freight_charges or 0), 2)


# Sequence counter for quotation numbers, stored in DB so it survives restarts
class Counter(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    value = db.Column(db.Integer, nullable=False)


def next_quote_number():
    counter = Counter.query.filter_by(name='quote_no').first()
    if not counter:
        counter = Counter(name='quote_no', value=260001)
        db.session.add(counter)
    else:
        counter.value += 1
    db.session.commit()
    return str(counter.value)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            flash("Admin access required.", "danger")
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return wrapper


def visible_documents_query():
    """Admins see all documents; staff see only their own."""
    if current_user.is_admin():
        return Document.query
    return Document.query.filter_by(created_by_id=current_user.id)


def visible_customers_query():
    if current_user.is_admin():
        return Customer.query
    return Customer.query.filter_by(created_by_id=current_user.id)


# ---------------------------------------------------------------------------
# AUTH ROUTES
# ---------------------------------------------------------------------------

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()
        if user and user.active and user.check_password(password):
            login_user(user)
            return redirect(url_for('dashboard'))
        flash("Invalid username or password.", "danger")
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


# ---------------------------------------------------------------------------
# DASHBOARD
# ---------------------------------------------------------------------------

@app.route('/')
@login_required
def dashboard():
    docs = visible_documents_query().all()
    total_amount = sum(d.total_amount or 0 for d in docs)
    pending_docs = [d for d in docs if d.status != 'delivered']
    pending_amount = sum(d.total_amount or 0 for d in pending_docs)
    delivered_count = len(docs) - len(pending_docs)

    today = date.today()
    overdue = [d for d in pending_docs if d.follow_up_date and d.follow_up_date < today]
    due_soon = [d for d in pending_docs if d.follow_up_date and 0 <= (d.follow_up_date - today).days <= 2]

    pending_sorted = sorted(
        pending_docs,
        key=lambda d: (d.follow_up_date or date.max)
    )[:10]

    return render_template(
        'dashboard.html',
        total_amount=total_amount,
        pending_amount=pending_amount,
        delivered_count=delivered_count,
        pending_count=len(pending_docs),
        overdue=overdue,
        due_soon=due_soon,
        pending_docs=pending_sorted,
        today=today,
    )


# ---------------------------------------------------------------------------
# DOCUMENTS (PI / QUOTATIONS)
# ---------------------------------------------------------------------------

@app.route('/documents')
@login_required
def documents():
    q = request.args.get('q', '').strip()
    status_filter = request.args.get('status', '')

    query = visible_documents_query()
    if status_filter:
        query = query.filter(Document.status == status_filter)
    docs = query.order_by(Document.created_at.desc()).all()

    if q:
        ql = q.lower()
        docs = [d for d in docs if ql in d.item_desc.lower()
                or ql in (d.customer.name.lower() if d.customer else '')
                or ql in d.quote_no.lower()]

    customers = visible_customers_query().order_by(Customer.name).all()
    items = Item.query.order_by(Item.name).all()

    return render_template(
        'documents.html',
        docs=docs, customers=customers, items=items,
        dispatch_units=DISPATCH_UNITS, uom_options=UOM_OPTIONS,
        q=q, status_filter=status_filter,
        today=date.today().isoformat(),
    )


@app.route('/documents/new', methods=['POST'])
@login_required
def new_document():
    customer_id = request.form.get('customer_id')
    item_desc = request.form.get('item_desc', '').strip()
    dispatch_from = request.form.get('dispatch_from')

    if not customer_id or not item_desc or not dispatch_from:
        flash("Customer, dispatch location, and item are required.", "danger")
        return redirect(url_for('documents'))

    doc = Document(
        doc_type=request.form.get('doc_type', 'PI'),
        quote_no=next_quote_number(),
        doc_date=date.today(),
        dispatch_from=dispatch_from,
        customer_id=int(customer_id),
        item_desc=item_desc,
        packaging=request.form.get('packaging', '').strip(),
        qty=float(request.form.get('qty') or 0),
        uom=request.form.get('uom'),
        price=float(request.form.get('price') or 0),
        gst_applied=request.form.get('gst_applied') == 'on',
        freight_charges=float(request.form.get('freight_charges') or 0),
        follow_up_date=_parse_date(request.form.get('follow_up_date')),
        status=request.form.get('status', 'pending'),
        notes=request.form.get('notes', '').strip(),
        created_by_id=current_user.id,
    )
    doc.recalc()
    db.session.add(doc)
    db.session.commit()
    flash(f"Document {doc.quote_no} created.", "success")
    return redirect(url_for('documents'))


@app.route('/documents/<int:doc_id>/update', methods=['POST'])
@login_required
def update_document(doc_id):
    doc = Document.query.get_or_404(doc_id)
    if not current_user.is_admin() and doc.created_by_id != current_user.id:
        flash("You can only edit your own documents.", "danger")
        return redirect(url_for('documents'))

    doc.status = request.form.get('status', doc.status)
    doc.follow_up_date = _parse_date(request.form.get('follow_up_date')) or doc.follow_up_date
    doc.notes = request.form.get('notes', doc.notes)
    db.session.commit()
    flash("Document updated.", "success")
    return redirect(url_for('documents'))


@app.route('/documents/<int:doc_id>/delete', methods=['POST'])
@login_required
def delete_document(doc_id):
    doc = Document.query.get_or_404(doc_id)
    if not current_user.is_admin() and doc.created_by_id != current_user.id:
        flash("You can only delete your own documents.", "danger")
        return redirect(url_for('documents'))
    db.session.delete(doc)
    db.session.commit()
    flash("Document deleted.", "success")
    return redirect(url_for('documents'))


@app.route('/documents/<int:doc_id>/pdf')
@login_required
def download_document_pdf(doc_id):
    doc = Document.query.get_or_404(doc_id)
    if not current_user.is_admin() and doc.created_by_id != current_user.id:
        flash("You can only download your own documents.", "danger")
        return redirect(url_for('documents'))

    from flask import send_file
    buffer = generate_pi_pdf(doc)
    filename = f"{doc.doc_type}_{doc.quote_no}_{doc.customer.name[:20].replace(' ', '_')}.pdf"
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename,
    )


def _parse_date(val):
    if not val:
        return None
    try:
        return datetime.strptime(val, '%Y-%m-%d').date()
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# CUSTOMERS  (+ bulk import)
# ---------------------------------------------------------------------------

@app.route('/customers')
@login_required
def customers():
    custs = visible_customers_query().order_by(Customer.name).all()
    return render_template('customers.html', customers=custs)


@app.route('/customers/new', methods=['POST'])
@login_required
def new_customer():
    name = request.form.get('name', '').strip()
    if not name:
        flash("Customer name is required.", "danger")
        return redirect(url_for('customers'))
    c = Customer(
        name=name,
        contact_person=request.form.get('contact_person', '').strip(),
        phone=request.form.get('phone', '').strip(),
        email=request.form.get('email', '').strip(),
        created_by_id=current_user.id,
    )
    db.session.add(c)
    db.session.commit()
    flash(f"Customer '{name}' added.", "success")
    return redirect(url_for('customers'))


@app.route('/customers/<int:cust_id>/delete', methods=['POST'])
@login_required
def delete_customer(cust_id):
    c = Customer.query.get_or_404(cust_id)
    if not current_user.is_admin() and c.created_by_id != current_user.id:
        flash("You can only delete your own customers.", "danger")
        return redirect(url_for('customers'))
    db.session.delete(c)
    db.session.commit()
    flash("Customer deleted.", "success")
    return redirect(url_for('customers'))


@app.route('/customers/bulk-import', methods=['POST'])
@login_required
def bulk_import_customers():
    file = request.files.get('file')
    if not file or file.filename == '':
        flash("Please choose a file to import.", "danger")
        return redirect(url_for('customers'))

    rows = _read_tabular_file(file)
    if rows is None:
        flash("Unsupported file type. Use .csv or .xlsx", "danger")
        return redirect(url_for('customers'))

    created = 0
    for row in rows:
        name = (row.get('name') or row.get('Name') or row.get('Customer Name') or '').strip()
        if not name:
            continue
        c = Customer(
            name=name,
            contact_person=(row.get('contact_person') or row.get('Contact Person') or '').strip(),
            phone=(row.get('phone') or row.get('Phone') or row.get('WhatsApp') or '').strip(),
            email=(row.get('email') or row.get('Email') or '').strip(),
            created_by_id=current_user.id,
        )
        db.session.add(c)
        created += 1
    db.session.commit()
    flash(f"Imported {created} customers.", "success")
    return redirect(url_for('customers'))


@app.route('/customers/template')
@login_required
def customer_template():
    return _csv_response(
        ['name', 'contact_person', 'phone', 'email'],
        [['Example Traders Pvt Ltd', 'Ramesh Kumar', '+919876543210', 'ramesh@example.com']],
        'customer_import_template.csv'
    )


# ---------------------------------------------------------------------------
# ITEMS  (+ bulk import)
# ---------------------------------------------------------------------------

@app.route('/items')
@login_required
def items():
    all_items = Item.query.order_by(Item.name).all()
    return render_template('items.html', items=all_items, uom_options=UOM_OPTIONS)


@app.route('/items/new', methods=['POST'])
@login_required
def new_item():
    name = request.form.get('name', '').strip()
    if not name:
        flash("Item name is required.", "danger")
        return redirect(url_for('items'))
    item = Item(
        name=name,
        default_uom=request.form.get('default_uom'),
        default_price=float(request.form.get('default_price') or 0),
        created_by_id=current_user.id,
    )
    db.session.add(item)
    db.session.commit()
    flash(f"Item '{name}' added.", "success")
    return redirect(url_for('items'))


@app.route('/items/<int:item_id>/delete', methods=['POST'])
@login_required
def delete_item(item_id):
    item = Item.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash("Item deleted.", "success")
    return redirect(url_for('items'))


@app.route('/items/bulk-import', methods=['POST'])
@login_required
def bulk_import_items():
    file = request.files.get('file')
    if not file or file.filename == '':
        flash("Please choose a file to import.", "danger")
        return redirect(url_for('items'))

    rows = _read_tabular_file(file)
    if rows is None:
        flash("Unsupported file type. Use .csv or .xlsx", "danger")
        return redirect(url_for('items'))

    created = 0
    for row in rows:
        name = (row.get('name') or row.get('Name') or row.get('Item') or '').strip()
        if not name:
            continue
        item = Item(
            name=name,
            default_uom=(row.get('uom') or row.get('UOM') or '').strip(),
            default_price=float(row.get('price') or row.get('Price') or 0 or 0),
            created_by_id=current_user.id,
        )
        db.session.add(item)
        created += 1
    db.session.commit()
    flash(f"Imported {created} items.", "success")
    return redirect(url_for('items'))


@app.route('/items/template')
@login_required
def item_template():
    return _csv_response(
        ['name', 'uom', 'price'],
        [['28 mm caps - K blue caps - 20 box', 'Nos', '0.32']],
        'item_import_template.csv'
    )


# ---------------------------------------------------------------------------
# ADMIN: USER MANAGEMENT
# ---------------------------------------------------------------------------

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    all_users = User.query.order_by(User.username).all()
    return render_template('admin_users.html', users=all_users)


@app.route('/admin/users/new', methods=['POST'])
@login_required
@admin_required
def admin_new_user():
    username = request.form.get('username', '').strip().lower()
    full_name = request.form.get('full_name', '').strip()
    password = request.form.get('password', '')
    role = request.form.get('role', 'staff')

    if not username or not password or not full_name:
        flash("Username, full name, and password are required.", "danger")
        return redirect(url_for('admin_users'))

    if User.query.filter_by(username=username).first():
        flash("That username already exists.", "danger")
        return redirect(url_for('admin_users'))

    user = User(username=username, full_name=full_name, role=role)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    flash(f"User '{username}' created.", "success")
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
@login_required
@admin_required
def admin_toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash("You cannot deactivate your own account.", "danger")
        return redirect(url_for('admin_users'))
    user.active = not user.active
    db.session.commit()
    flash(f"User '{user.username}' {'activated' if user.active else 'deactivated'}.", "success")
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
@admin_required
def admin_reset_password(user_id):
    user = User.query.get_or_404(user_id)
    new_password = request.form.get('new_password', '')
    if len(new_password) < 6:
        flash("Password must be at least 6 characters.", "danger")
        return redirect(url_for('admin_users'))
    user.set_password(new_password)
    db.session.commit()
    flash(f"Password reset for '{user.username}'.", "success")
    return redirect(url_for('admin_users'))


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def _read_tabular_file(file_storage):
    filename = file_storage.filename.lower()
    if filename.endswith('.csv'):
        stream = io.StringIO(file_storage.stream.read().decode('utf-8-sig'))
        reader = csv.DictReader(stream)
        return list(reader)
    elif filename.endswith('.xlsx'):
        wb = load_workbook(file_storage, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h else '' for h in next(rows_iter)]
        except StopIteration:
            return []
        result = []
        for row in rows_iter:
            row_dict = {header[i]: (row[i] if i < len(row) else None) for i in range(len(header))}
            cleaned = {k: ('' if v is None else str(v)) for k, v in row_dict.items()}
            result.append(cleaned)
        return result
    return None


def _csv_response(headers, rows, filename):
    from flask import Response
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


# ---------------------------------------------------------------------------
# CLI: create initial admin user + tables
# ---------------------------------------------------------------------------

@app.cli.command('init-db')
def init_db():
    """Create tables and a default admin user (run once)."""
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', full_name='Administrator', role='admin')
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        print("Created default admin user -> username: admin / password: admin123")
        print("IMPORTANT: log in and change this password immediately.")
    else:
        print("Admin user already exists.")


with app.app_context():
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        _admin = User(username='admin', full_name='Administrator', role='admin')
        _admin.set_password('admin123')
        db.session.add(_admin)
        db.session.commit()


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
